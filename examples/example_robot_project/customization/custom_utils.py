# -----------------------------------------------------------------------------
# Copyright 2017 Aleksey Litvinov litvinov.aleks@gmail.com
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# -----------------------------------------------------------------------------

# -----------------------------------------------------------------------------
# This file will contain utilities function for specific project.
# -----------------------------------------------------------------------------
from openpyxl import load_workbook


# <CUSTOMIZATION>
def is_script(root, full_file_name):
    """
    Function decides if current file a script according to its root and full name
    :param root: path to the file
    :param full_file_name: name of file with extension (if present)
    :return: boolean value, if file is a script
    """
    is_script_result = False
    if '.' in full_file_name:
        file_name, extension = full_file_name.rsplit('.', 1)
        if extension in ["xlsx", "xls"] and file_name.startswith("RobotScenario_"):
            is_script_result = True
    return is_script_result
# </CUSTOMIZATION>


# <CUSTOMIZATION>
def get_configuration_parameters(config):
    """
    Creates dictionary with parameters taken from/based on configuration file object
    :param config: object for configuration file
    :return: dictionary with custom parameters taken from/based on configuration file object
    """
    kwargs = {}
    kwargs["scenario_sheet"] = config["CUSTOM_PARAMETERS"]["Scenario_sheet"]
    kwargs["expected_headers"] = [feature.strip() for feature in config["CUSTOM_PARAMETERS"]["Expected_headers"].split(',')]
    kwargs["commands"] = [feature.strip() for feature in config["CUSTOM_PARAMETERS"]["Commands"].split(',')]
    return kwargs
# </CUSTOMIZATION>


# <CUSTOMIZATION>
def get_script(script_path):
    """
    Returns script object by its path (for example for file-like object call open(script_path)).
    File preprocessing could be implemented here.
    :param script_path: path to script file
    :return: script object
    """
    wb = load_workbook(filename=script_path)
    return wb
# </CUSTOMIZATION>


# <CUSTOMIZATION>
def close_script(script):
    """
    Perform action on script object close (for example, if script is file-like object call script.close())
    :param script: script object
    :return: None
    """
    pass
# </CUSTOMIZATION>


def get_rows_list(worksheet):
    """
    Get worksheet data as a list of rows. Every row presented as a list of cell string value.
    :param worksheet: worksheet object
    :return: list of Excel worksheet rows as a list of string values (empty cell will return 'None' value)
    """
    rows = []
    for row in worksheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(str(cell.value))
        rows.append(row_data)
    return rows


def get_scenario_header(worksheet):
    """
    Get first row of worksheet as a list.
    :param worksheet: worksheet object
    :return: list of first Excel worksheet row as a list of string values (empty cell will return 'None' value)
    """
    return get_rows_list(worksheet)[0]


def get_scenario_data(worksheet):
    """
    Get rows of worksheet as a list, except first row (header).
    :param worksheet: worksheet object
    :return: list of Excel worksheet rows (except first row (header)) as a list of string values
    (empty cell will return 'None' value)
    """
    return get_rows_list(worksheet)[1:]


def is_int(string_to_check):
    """
    Checks if string is an integer value.
    :param string_to_check: String to check
    :return: True if string is an integer, False otherwise
    """
    try:
        int(string_to_check)
        result = True
    except ValueError:
        result = False
    return result
